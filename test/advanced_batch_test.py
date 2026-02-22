#!/usr/bin/env python3
"""
Advanced Batch Refactoring Test Script - Version Finale
- Une seule ligne par test (pas de duplication)
- Colonnes dynamiques : Durée_RenameAgent, Temp_RenameAgent, Durée_ImportAgent, etc.
- Durées RÉELLES extraites du workflow LangGraph
- Statut basé sur TestAgent
"""

import os
import sys
import time
import json
import csv
import argparse
from datetime import datetime
from pathlib import Path
import traceback

# Ajouter le répertoire parent au path
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
sys.path.insert(0, parent_dir)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from core.ollama_llm_client import OllamaLLMClient
from core.langgraph_orchestrator import LangGraphOrchestrator


class AdvancedBatchTester:
    """Testeur batch avec support LangGraph"""
    
    def __init__(self, config_file="test_config.json"):
        """Initialise le testeur"""
        self.load_config(config_file)
        self.output_dir.mkdir(exist_ok=True)
        
        print("🔄 Initialisation du système...")
        self.llm_client = OllamaLLMClient(model_name=self.config.get('model', 'mistral:latest'))
        self.orchestrator = LangGraphOrchestrator(self.llm_client)
        
        self.results = []
        self.errors = []
        
        print(f"✅ Système initialisé avec {len(self.test_configurations)} configurations")
    
    def load_config(self, config_file):
        """Charge la configuration"""
        print(f"📄 Chargement: {config_file}")
        
        with open(config_file, 'r', encoding='utf-8') as f:
            self.config = json.load(f)
        
        self.input_dir = Path(self.config.get('input_directory', 'bad_codes'))
        self.output_dir = Path(self.config.get('output_directory', 'test_results'))
        self.test_configurations = self.config.get('test_configurations', [])
        self.output_options = self.config.get('output_options', {})
        
        print(f"   📁 Entrée: {self.input_dir}")
        print(f"   📁 Sortie: {self.output_dir}")
    
    def get_test_files(self, pattern="bad_code*.py"):
        """Récupère les fichiers"""
        files = sorted(self.input_dir.glob(pattern))
        print(f"📁 {len(files)} fichiers trouvés")
        return files
    
    def format_duration(self, seconds):
        """Formate une durée"""
        if seconds < 1:
            return f"{seconds*1000:.0f}ms"
        elif seconds < 60:
            return f"{seconds:.2f}s"
        else:
            minutes = int(seconds // 60)
            secs = seconds % 60
            return f"{minutes}m {secs:.2f}s"
    
    def run_test(self, file_path, test_config):
        """Execute un test avec workflow LangGraph"""
        print(f"\n🧪 [{test_config['id']}] {file_path.name}")
        
        test_start_time = time.time()
        
        try:
            # Lire le code
            with open(file_path, 'r', encoding='utf-8') as f:
                code = f.read()
            
            original_lines = len(code.split('\n'))
            original_chars = len(code)
            
            # Préparer les agents
            selected_agents = [agent['name'] for agent in test_config['agents']]
            print(f"  🚀 Workflow LangGraph avec {len(selected_agents)} agents...")
            
            # Configurer les températures personnalisées
            temperature_override = {}
            for agent_config in test_config['agents']:
                agent_name = agent_config['name']
                agent_temp = agent_config.get('temperature')
                if agent_temp is not None:
                    temperature_override[agent_name] = agent_temp
            
            print(f"   🌡️  Températures configurées: {temperature_override}")
            
            # Exécuter le workflow
            workflow_start = time.time()
            
            workflow_result = self.orchestrator.run_workflow(
                code=code,
                language="Python",
                selected_agents=selected_agents,
                auto_patch=self.output_options.get("auto_patch", True),
                auto_test=self.output_options.get("auto_test", True),
                temperature_override=temperature_override  # ⭐ IMPORTANT
            )
            
            workflow_duration = time.time() - workflow_start

            if not workflow_result.get("success"):
                raise Exception(workflow_result.get("error", "Workflow failed"))

            # Code final
            final_code = workflow_result.get("refactored_code", code)
            
            # ========== EXTRACTION DES DONNÉES PAR AGENT ==========
            agent_details_dict = {}  # Dictionnaire pour stocker les détails par agent
            total_issues_found = 0
            test_status = "SUCCESS"  # Par défaut
            
            print(f"\n  ⏱️  Durées individuelles des agents:")
            
            # Parcourir les résultats du workflow
            for agent_result in workflow_result.get("agent_results", []):
                agent_name = agent_result.get("name")
                
                # ⭐ Récupérer la VRAIE durée depuis le workflow
                agent_duration = agent_result.get("duration", 0)
                
                # Analyse et statut
                analysis = agent_result.get("analysis", [])
                issues = len(analysis)
                total_issues_found += issues
                
                agent_status = agent_result.get("status", "SUCCESS")
                temp_used = agent_result.get("temperature_used")
                
                # Afficher
                status_icon = "✅" if agent_status == "SUCCESS" else "❌"
                print(f"     {status_icon} {agent_name:20s}: {self.format_duration(agent_duration):>8s} | {issues:2d} problèmes | Temp: {temp_used} | {agent_status}")
                
                # ⭐ Vérifier TestAgent pour le statut global
                if agent_name == "TestAgent":
                    if agent_status != "SUCCESS":
                        test_status = "FAILED"
                        print(f"        ⚠️  TestAgent a échoué → Test marqué FAILED")
                
                # ⭐ Stocker dans le dictionnaire (une entrée par agent)
                agent_details_dict[agent_name] = {
                    "duration": agent_duration,
                    "duration_formatted": self.format_duration(agent_duration),
                    "temperature": temp_used,
                    "issues_found": issues,
                    "status": agent_status
                }
            
            total_test_duration = time.time() - test_start_time
            final_lines = len(final_code.split('\n'))
            final_chars = len(final_code)
            
            # Afficher résumé
            print(f"\n  ⏱️  Durée totale: {self.format_duration(total_test_duration)}")
            print(f"  📊  {total_issues_found} problèmes | {abs(final_lines - original_lines)} lignes changées")
            
            # Sauvegarder
            if self.output_options.get('save_refactored_code', True):
                output_file = self.output_dir / f"{file_path.stem}_{test_config['id']}_refactored.py"
                with open(output_file, 'w', encoding='utf-8') as f:
                    f.write(final_code)
            else:
                output_file = None
            
            # ⭐ Résultat final avec dictionnaire d'agents
            result = {
                'file': file_path.name,
                'test_config': test_config['name'],
                'test_id': test_config['id'],
                'status': test_status,  # ⭐ Basé sur TestAgent
                'total_duration': total_test_duration,
                'workflow_duration': workflow_duration,
                'agents_used': len(selected_agents),
                'total_issues': total_issues_found,
                'original_lines': original_lines,
                'final_lines': final_lines,
                'lines_changed': abs(final_lines - original_lines),
                'original_chars': original_chars,
                'final_chars': final_chars,
                'chars_changed': abs(final_chars - original_chars),
                'reduction_percent': ((original_chars - final_chars) / original_chars * 100) if original_chars > 0 else 0,
                'agent_details_dict': agent_details_dict,  # ⭐ Dictionnaire au lieu de liste
                'output_file': str(output_file) if output_file else None,
                'timestamp': datetime.now().isoformat()
            }
            
            status_icon = "✅" if test_status == "SUCCESS" else "❌"
            print(f"\n  {status_icon} Statut final: {test_status}")
            
            return result
            
        except Exception as e:
            total_duration = time.time() - test_start_time
            error_result = {
                'file': file_path.name,
                'test_config': test_config['name'],
                'test_id': test_config['id'],
                'status': f'FAILED: {str(e)[:100]}',
                'total_duration': total_duration,
                'error_trace': traceback.format_exc(),
                'timestamp': datetime.now().isoformat()
            }
            
            if self.output_options.get('save_error_logs', True):
                error_file = self.output_dir / f"error_{file_path.stem}_{test_config['id']}.log"
                with open(error_file, 'w', encoding='utf-8') as f:
                    f.write(f"Error: {test_config['id']}\n")
                    f.write(f"File: {file_path}\n")
                    f.write(f"Time: {error_result['timestamp']}\n\n")
                    f.write(error_result['error_trace'])
            
            print(f"  ❌ Échec: {str(e)[:50]}")
            return error_result
    
    def run_all_tests(self, file_pattern="bad_code*.py", test_ids=None):
        """Execute tous les tests"""
        files = self.get_test_files(file_pattern)
        
        if not files:
            print("❌ Aucun fichier!")
            return
        
        configs = self.test_configurations
        if test_ids:
            configs = [c for c in configs if c['id'] in test_ids]
            print(f"🔍 {len(configs)} configurations sélectionnées")
        
        total_tests = len(files) * len(configs)
        current_test = 0
        
        print(f"\n🚀 {total_tests} tests")
        print(f"   ({len(files)} fichiers × {len(configs)} configs)\n")
        
        start_time = time.time()
        
        for file_path in files:
            for config in configs:
                current_test += 1
                print(f"\n{'='*70}")
                print(f"Test {current_test}/{total_tests}")
                
                result = self.run_test(file_path, config)
                self.results.append(result)
                
                if result['status'] not in ['SUCCESS']:
                    self.errors.append(result)
        
        total_time = time.time() - start_time
        
        print(f"\n{'='*70}")
        print(f"✅ Terminé en {self.format_duration(total_time)}")
        print(f"   - Succès: {len([r for r in self.results if r['status'] == 'SUCCESS'])}")
        print(f"   - Échecs: {len(self.errors)}")
    
    def export_to_excel(self, filename=None):
        """Export Excel avec colonnes dynamiques par agent"""
        
        if not self.results:
            print("❌ Aucun résultat")
            return
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"batch_results_{timestamp}.xlsx"
        
        print(f"\n📊 Export Excel: {filename}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Résultats"
        
        # ========== TROUVER TOUS LES AGENTS UTILISÉS ==========
        all_agents = set()
        for result in self.results:
            if 'agent_details_dict' in result:
                all_agents.update(result['agent_details_dict'].keys())
        
        all_agents = sorted(list(all_agents))  # Ordre alphabétique
        
        print(f"   📋 Agents détectés: {', '.join(all_agents)}")
        
        # ========== CONSTRUIRE LES EN-TÊTES DYNAMIQUES ==========
        headers_base = [
            'Fichier', 'Configuration', 'Test_ID', 'Statut_Test',
            'Durée_Totale', 'Durée_Workflow', 'Nb_Agents',
            'Problèmes_Totaux', 'Lignes_Avant', 'Lignes_Après', 
            'Δ_Lignes', '% Réduction'
        ]
        
        # Ajouter des colonnes pour chaque agent
        headers_agents = []
        for agent in all_agents:
            headers_agents.extend([
                f'Durée_{agent}',
                f'Temp_{agent}',
                f'Problèmes_{agent}',
                f'Statut_{agent}'
            ])
        
        headers = headers_base + headers_agents + ['Timestamp']
        
        ws.append(headers)
        
        # Formater en-tête
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # ========== AJOUTER LES DONNÉES (UNE LIGNE PAR TEST) ==========
        for result in self.results:
            if 'agent_details_dict' in result:
                # Données de base
                row = [
                    result['file'],
                    result['test_config'],
                    result['test_id'],
                    result['status'],
                    self.format_duration(result.get('total_duration', 0)),
                    self.format_duration(result.get('workflow_duration', 0)),
                    result.get('agents_used', 0),
                    result.get('total_issues', 0),
                    result.get('original_lines', 0),
                    result.get('final_lines', 0),
                    result.get('lines_changed', 0),
                    f"{result.get('reduction_percent', 0):.1f}%"
                ]
                
                # ⭐ Ajouter les données de chaque agent
                agent_details = result.get('agent_details_dict', {})
                for agent in all_agents:
                    if agent in agent_details:
                        details = agent_details[agent]
                        row.extend([
                            details['duration_formatted'],  # Durée formatée
                            details['temperature'] if details['temperature'] is not None else 'N/A',
                            details['issues_found'],
                            details['status']
                        ])
                    else:
                        # Agent pas utilisé dans ce test
                        row.extend(['', '', '', ''])
                
                row.append(result.get('timestamp', ''))
                ws.append(row)
            else:
                # Test échoué
                row = [
                    result['file'], result['test_config'], result['test_id'],
                    result['status'], self.format_duration(result.get('total_duration', 0)),
                    '', '', '', '', '', '', ''
                ]
                # Remplir colonnes agents vides
                row.extend([''] * len(headers_agents))
                row.append(result.get('timestamp', ''))
                ws.append(row)
        
        # Ajuster largeurs
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Feuille Statistiques
        ws_stats = wb.create_sheet("Statistiques")
        
        successful = [r for r in self.results if r['status'] == 'SUCCESS']
        failed = [r for r in self.results if r['status'] == 'FAILED']
        total = len(self.results)
        
        stats = [
            ['Métrique', 'Valeur'],
            ['Tests Totaux', total],
            ['Réussis', len(successful)],
            ['Échoués', len(failed)],
            ['Taux Réussite', f"{(len(successful)/total*100):.1f}%" if total > 0 else "0%"],
            [''],
            ['Durée Totale', self.format_duration(sum([r.get('total_duration', 0) for r in self.results]))],
            ['Durée Moyenne', self.format_duration(sum([r.get('total_duration', 0) for r in self.results])/total) if total > 0 else "0s"],
            [''],
            ['Problèmes Totaux', sum([r.get('total_issues', 0) for r in successful])],
            ['Problèmes Moyens', f"{sum([r.get('total_issues', 0) for r in successful])/len(successful):.1f}" if successful else "0"],
        ]
        
        for row in stats:
            ws_stats.append(row)
        
        for row in range(1, len(stats) + 1):
            cell_a = ws_stats.cell(row=row, column=1)
            cell_b = ws_stats.cell(row=row, column=2)
            
            if row == 1:
                cell_a.font = Font(bold=True, color='FFFFFF')
                cell_b.font = Font(bold=True, color='FFFFFF')
                cell_a.fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')
                cell_b.fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')
            else:
                cell_a.font = Font(bold=True)
        
        ws_stats.column_dimensions['A'].width = 30
        ws_stats.column_dimensions['B'].width = 20
        
        output_path = self.output_dir / filename
        wb.save(output_path)
        
        print(f"✅ Excel: {output_path}")
        print(f"   - Une ligne par test avec colonnes par agent")
        print(f"   - Agents: {', '.join(all_agents)}")
        return output_path
    
    def export_to_csv(self, filename=None):
        """Export CSV avec colonnes dynamiques"""
        if not self.results:
            print("❌ Aucun résultat")
            return
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"batch_results_{timestamp}.csv"
        
        output_path = self.output_dir / filename
        
        print(f"\n📊 Export CSV: {filename}")
        
        # Trouver tous les agents
        all_agents = set()
        for result in self.results:
            if 'agent_details_dict' in result:
                all_agents.update(result['agent_details_dict'].keys())
        
        all_agents = sorted(list(all_agents))
        
        # Construire les en-têtes
        headers = [
            'Fichier', 'Configuration', 'Test_ID', 'Statut',
            'Durée_Totale_Sec', 'Durée_Totale', 
            'Durée_Workflow_Sec', 'Durée_Workflow',
            'Problèmes_Totaux', 'Lignes_Avant', 'Lignes_Après', 
            'Lignes_Changées', 'Réduction_%'
        ]
        
        # Ajouter colonnes par agent
        for agent in all_agents:
            headers.extend([
                f'Durée_{agent}_Sec',
                f'Durée_{agent}',
                f'Temp_{agent}',
                f'Problèmes_{agent}',
                f'Statut_{agent}'
            ])
        
        headers.append('Timestamp')
        
        # Écrire CSV
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=headers)
            writer.writeheader()
            
            for result in self.results:
                if 'agent_details_dict' in result:
                    row = {
                        'Fichier': result['file'],
                        'Configuration': result['test_config'],
                        'Test_ID': result['test_id'],
                        'Statut': result['status'],
                        'Durée_Totale_Sec': result.get('total_duration', 0),
                        'Durée_Totale': self.format_duration(result.get('total_duration', 0)),
                        'Durée_Workflow_Sec': result.get('workflow_duration', 0),
                        'Durée_Workflow': self.format_duration(result.get('workflow_duration', 0)),
                        'Problèmes_Totaux': result.get('total_issues', 0),
                        'Lignes_Avant': result.get('original_lines', 0),
                        'Lignes_Après': result.get('final_lines', 0),
                        'Lignes_Changées': result.get('lines_changed', 0),
                        'Réduction_%': f"{result.get('reduction_percent', 0):.2f}",
                        'Timestamp': result.get('timestamp', '')
                    }
                    
                    # Ajouter données agents
                    agent_details = result.get('agent_details_dict', {})
                    for agent in all_agents:
                        if agent in agent_details:
                            details = agent_details[agent]
                            row[f'Durée_{agent}_Sec'] = details['duration']
                            row[f'Durée_{agent}'] = details['duration_formatted']
                            row[f'Temp_{agent}'] = details['temperature'] if details['temperature'] is not None else 'N/A'
                            row[f'Problèmes_{agent}'] = details['issues_found']
                            row[f'Statut_{agent}'] = details['status']
                        else:
                            row[f'Durée_{agent}_Sec'] = ''
                            row[f'Durée_{agent}'] = ''
                            row[f'Temp_{agent}'] = ''
                            row[f'Problèmes_{agent}'] = ''
                            row[f'Statut_{agent}'] = ''
                    
                    writer.writerow(row)
        
        print(f"✅ CSV: {output_path}")
        print(f"   - {len(headers)} colonnes")
        return output_path


def main():
    """Fonction principale"""
    
    parser = argparse.ArgumentParser(description='Batch Refactoring Test')
    parser.add_argument('--config', default='test_config.json', help='Config JSON')
    parser.add_argument('--pattern', default='bad_code*.py', help='Fichiers')
    parser.add_argument('--tests', nargs='+', help='IDs des tests')
    parser.add_argument('--excel', default=None, help='Nom Excel')
    parser.add_argument('--csv', default=None, help='Nom CSV')
    
    args = parser.parse_args()
    
    tester = AdvancedBatchTester(config_file=args.config)
    tester.run_all_tests(file_pattern=args.pattern, test_ids=args.tests)
    
    if tester.output_options.get('generate_excel_report', True):
        tester.export_to_excel(filename=args.excel)
    
    tester.export_to_csv(filename=args.csv)
    
    print("\n🎉 Terminé!")


if __name__ == "__main__":
    main()